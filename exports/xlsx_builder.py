"""Excel export builder.

Implements the contract in CONTRACTS.md §5.4. Phase 5 adds the Scorecard
sheet; ``scenario`` remains deferred (`NotImplementedError`) until Phase 7.

Pages call this through ``build(modules, filters)`` and pipe the returned
``bytes`` straight into ``st.download_button`` (rule §3.4).
"""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO

import duckdb
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

import config
from data import query
from data.query import Filters

# ---------------------------------------------------------------------------
# Style tokens (TKO — SPEC §6.2 formatting rules)
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill(
    "solid", fgColor=config.TKO["colors"]["cold_slate"].lstrip("#")
)
_HEADER_FONT = Font(
    name=config.TKO["fonts"]["body"],
    bold=True,
    color=config.TKO["colors"]["ice_white"].lstrip("#"),
)
_TITLE_FONT = Font(
    name=config.TKO["fonts"]["display"], size=16, bold=True,
    color=config.TKO["colors"]["void"].lstrip("#"),
)
_SECTION_FONT = Font(
    name=config.TKO["fonts"]["display"], size=12, bold=True,
    color=config.TKO["colors"]["void"].lstrip("#"),
)
_LABEL_FONT = Font(name=config.TKO["fonts"]["body"], bold=True)
_BORDER = Border(
    bottom=Side(style="thin", color=config.TKO["colors"]["cold_slate"].lstrip("#"))
)
_RAG_FILLS = {
    rag: PatternFill("solid", fgColor=color.lstrip("#"))
    for rag, color in config.RAG_COLORS.items()
}

# Number formats
_FMT_INT = "#,##0"
_FMT_PCT_1DP = '0.0"%"'        # values are already in % units (e.g. 12.5)
_FMT_PCT_SIGNED = '+0.0"%";-0.0"%";0.0"%"'
_FMT_DATE = "yyyy-mm"


# ---------------------------------------------------------------------------
# Public entry
# ---------------------------------------------------------------------------


_SUPPORTED_MODULES = {"demand", "supply", "financial", "scorecard", "scenario"}

# Default scenario adjustments — used when the caller didn't pass one but
# 'scenario' is in modules (e.g. bundled export from a page that didn't touch
# the sliders yet). SPEC §5.5 — % volume adjustment vs consensus.
_DEFAULT_SCENARIO_ADJUSTMENTS = {"base": 0.0, "upside": 5.0, "downside": -5.0}


@st.cache_data(ttl=600, show_spinner=False)
def build(
    modules: list[str],
    filters: Filters,
    scenario_adjustments: dict | None = None,
) -> tuple[bytes, str]:
    """Render the requested modules into a workbook.

    Phase 5 supports ``demand``, ``supply``, ``financial``, and ``scorecard``.
    ``scenario`` raises ``NotImplementedError`` until Phase 7.

    Cached so identical (modules, filters) on repeated page renders return the
    same bytes without re-running the workbook construction. Upload page clears
    ``st.cache_data`` on every successful load (rule §3.3), so a fresh upload
    invalidates this naturally.
    """
    unsupported = [m for m in modules if m not in _SUPPORTED_MODULES]
    if unsupported:
        raise ValueError(
            f"xlsx_builder: unknown modules {unsupported}. "
            f"Supported: {sorted(_SUPPORTED_MODULES)}."
        )
    if not modules:
        raise ValueError("modules must include at least one of "
                         f"{sorted(_SUPPORTED_MODULES)}")

    cycle_month = _resolve_cycle_month(filters)

    wb = Workbook()
    # openpyxl creates an initial blank sheet — reuse it for Cover.
    cover = wb.active
    cover.title = "Cover"
    _write_cover(cover, modules, filters, cycle_month)

    # Stable sheet ordering regardless of caller's module list order.
    # Scorecard sits right after Cover (SPEC §6.2 — "always present").
    if "scorecard" in modules:
        _write_scorecard_sheet(wb.create_sheet("Scorecard"), filters)
    if "demand" in modules:
        _write_demand_sheet(wb.create_sheet("Demand"), filters)
    if "supply" in modules:
        _write_supply_sheet(wb.create_sheet("Supply"), filters)
    if "financial" in modules:
        _write_financial_sheet(wb.create_sheet("Financial"), filters)
    if "scenario" in modules:
        _write_scenario_sheet(
            wb.create_sheet("Scenario"),
            filters,
            scenario_adjustments or _DEFAULT_SCENARIO_ADJUSTMENTS,
        )

    buf = BytesIO()
    wb.save(buf)
    filename = f"S&OP_Report_{cycle_month}.xlsx"
    return buf.getvalue(), filename


# ---------------------------------------------------------------------------
# Cycle-month resolution (CONTRACTS §5.4 — same order as pptx_builder)
# ---------------------------------------------------------------------------


def _resolve_cycle_month(filters: Filters) -> str:
    """Return ``YYYY-MM`` per the contract's resolution order."""
    dv = query.current_data_version("demand")
    if dv is not None:
        # Direct DuckDB read is permitted inside exports/ (rule §3.6).
        c = duckdb.connect(config.DUCKDB_PATH, read_only=True)
        try:
            row = c.execute(
                "SELECT MAX(period_date) FROM fact_demand WHERE data_version = ?",
                [dv],
            ).fetchone()
        finally:
            c.close()
        if row and row[0] is not None:
            return pd.Timestamp(row[0]).strftime("%Y-%m")
    if filters.period_to is not None:
        return pd.Timestamp(filters.period_to).strftime("%Y-%m")
    return pd.Timestamp(date.today()).strftime("%Y-%m")


# ---------------------------------------------------------------------------
# Cover sheet
# ---------------------------------------------------------------------------


def _write_cover(
    ws: Worksheet, modules: list[str], f: Filters, cycle_month: str
) -> None:
    ws["A1"] = "S&OP Hub — Report"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A1:D1")

    rows = [
        ("Cycle month", cycle_month),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Period window", f"{f.period_from:%Y-%m} → {f.period_to:%Y-%m}"),
        ("Modules", ", ".join(modules)),
        ("Demand data version",    _version_label("demand")),
        ("Supply data version",    _version_label("supply")),
        ("Financial data version", _version_label("financial")),
        ("Currency", config.CURRENCY),
        ("", ""),
        ("Active filters", ""),
        ("  Brands",     _filter_label(f.brands)),
        ("  Categories", _filter_label(f.categories)),
        ("  Channels",   _filter_label(f.channels)),
        ("  Regions",    _filter_label(f.regions)),
        ("  SKUs",       _filter_label(f.skus)),
    ]
    for i, (label, value) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label).font = _LABEL_FONT
        ws.cell(row=i, column=2, value=value)

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 60
    ws.freeze_panes = "A2"


def _version_label(domain: str) -> str:
    v = query.current_data_version(domain)
    return f"v{v}" if v else "—"


def _filter_label(values: list[str]) -> str:
    return ", ".join(values) if values else "All"


# ---------------------------------------------------------------------------
# Scorecard sheet — tier-1 KPI table with RAG conditional formatting
# (SPEC §5.4 / §6.2 — always present in bundled exports)
# ---------------------------------------------------------------------------


def _write_scorecard_sheet(ws: Worksheet, f: Filters) -> None:
    ws["A1"] = "KPI Scorecard"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A1:D1")
    ws.cell(
        row=2, column=1,
        value=f"Window {f.period_from:%Y-%m} → {f.period_to:%Y-%m}",
    )

    df = query.scorecard_all(f)

    headers = ["KPI", "Value", "RAG", "Thresholds"]
    header_row = 4
    for i, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=i, value=h)
    _apply_header_row(ws, header_row, len(headers))

    if df.empty:
        ws.cell(row=header_row + 1, column=1, value="(no data)")
    else:
        for i, r in df.iterrows():
            row_idx = header_row + 1 + i
            ws.cell(row=row_idx, column=1, value=str(r["kpi_name"])).font = _LABEL_FONT

            v = r["value"]
            unit = r["unit"]
            if v is None or pd.isna(v):
                ws.cell(row=row_idx, column=2, value=None)
            elif unit == "days":
                cell = ws.cell(row=row_idx, column=2, value=float(v))
                cell.number_format = '0.0" days"'
            elif unit == "pp":
                cell = ws.cell(row=row_idx, column=2, value=float(v))
                cell.number_format = '+0.00"pp";-0.00"pp";0.00"pp"'
            elif r["kpi_name"] == "Forecast Bias":
                cell = ws.cell(row=row_idx, column=2, value=float(v))
                cell.number_format = _FMT_PCT_SIGNED
            else:  # % (FA, DOS-NA never lands here, Fill Rate, Revenue achievement)
                cell = ws.cell(row=row_idx, column=2, value=float(v))
                cell.number_format = _FMT_PCT_1DP

            rag = r["rag"] if pd.notna(r["rag"]) else "red"
            _apply_rag_cell(ws, row_idx, 3, rag)

            ws.cell(row=row_idx, column=4, value=str(r["threshold_explainer"]))

    for col, width in zip("ABCD", (28, 18, 10, 56)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A5"


# ---------------------------------------------------------------------------
# Demand sheet — summary on top, source tables stacked underneath
# (SPEC §6.2: one sheet per module, blank rows between sections)
# ---------------------------------------------------------------------------


def _write_demand_sheet(ws: Worksheet, f: Filters) -> None:
    ws["A1"] = "Demand Review"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A1:F1")
    ws.cell(row=2, column=1, value=f"Window {f.period_from:%Y-%m} → {f.period_to:%Y-%m}")

    row = 4
    row = _write_kpi_block(ws, row, f)
    row = _write_volume_bridge(ws, row + 2, f)
    row = _write_fcst_vs_actuals(ws, row + 2, f)
    row = _write_channel_mix(ws, row + 2, f)
    row = _write_mape_by_sku(ws, row + 2, f)

    for col, width in zip("ABCDEFGH", (18, 22, 22, 22, 22, 18, 18, 18)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "B2"   # row 1 + col A pinned (SPEC §6.2)


def _section_header(ws: Worksheet, row: int, title: str) -> int:
    ws.cell(row=row, column=1, value=title).font = _SECTION_FONT
    return row + 1


def _apply_header_row(ws: Worksheet, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = _BORDER


def _write_kpi_block(ws: Worksheet, row: int, f: Filters) -> int:
    row = _section_header(ws, row, "Headline KPIs")
    headers = ["Forecast Accuracy", "MAPE", "Bias", "Volume (cases)"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=row, column=i, value=h)
    _apply_header_row(ws, row, len(headers))
    row += 1

    k = query.demand_kpis(f)
    values = [k.fa_pct, k.mape, k.bias, k.volume_total]
    formats = [_FMT_PCT_1DP, _FMT_PCT_1DP, _FMT_PCT_SIGNED, _FMT_INT]
    for i, (v, fmt) in enumerate(zip(values, formats), start=1):
        if v is None or pd.isna(v):
            ws.cell(row=row, column=i, value=None)  # empty cell on no data
        else:
            cell = ws.cell(row=row, column=i, value=float(v))
            cell.number_format = fmt
    return row


def _write_volume_bridge(ws: Worksheet, row: int, f: Filters) -> int:
    row = _section_header(ws, row, "Volume Bridge")
    df = query.demand_volume_bridge(f)
    headers = ["Stage", "Value (cases)"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=row, column=i, value=h)
    _apply_header_row(ws, row, len(headers))
    row += 1
    if df.empty:
        ws.cell(row=row, column=1, value="(no data)")
        return row
    for _, r in df.iterrows():
        ws.cell(row=row, column=1, value=r["stage"])
        cell = ws.cell(row=row, column=2, value=float(r["value"]))
        cell.number_format = _FMT_INT
        row += 1
    return row - 1


def _write_fcst_vs_actuals(ws: Worksheet, row: int, f: Filters) -> int:
    row = _section_header(ws, row, "Forecast vs Actuals (monthly)")
    df = query.demand_fcst_vs_actuals(f)
    headers = ["Period", "Statistical Fcst", "Consensus Fcst", "Actuals"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=row, column=i, value=h)
    _apply_header_row(ws, row, len(headers))
    row += 1
    if df.empty:
        ws.cell(row=row, column=1, value="(no data)")
        return row
    for _, r in df.iterrows():
        period = pd.Timestamp(r["period_date"]).date()
        ws.cell(row=row, column=1, value=period).number_format = _FMT_DATE
        for i, col in enumerate(
            ("statistical_fcst", "consensus_fcst", "actuals"), start=2
        ):
            v = r[col]
            cell = ws.cell(
                row=row, column=i,
                value=float(v) if pd.notna(v) else None,
            )
            cell.number_format = _FMT_INT
        row += 1
    return row - 1


def _write_channel_mix(ws: Worksheet, row: int, f: Filters) -> int:
    row = _section_header(ws, row, "Channel Mix (share by period)")
    df = query.demand_channel_mix(f)
    headers = ["Period", "Channel Code", "Channel", "Volume Share %"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=row, column=i, value=h)
    _apply_header_row(ws, row, len(headers))
    row += 1
    if df.empty:
        ws.cell(row=row, column=1, value="(no data)")
        return row
    for _, r in df.iterrows():
        period = pd.Timestamp(r["period_date"]).date()
        ws.cell(row=row, column=1, value=period).number_format = _FMT_DATE
        ws.cell(row=row, column=2, value=r["channel_code"])
        ws.cell(row=row, column=3, value=r["channel_name"])
        cell = ws.cell(row=row, column=4, value=float(r["volume_share"]))
        cell.number_format = _FMT_PCT_1DP
        row += 1
    return row - 1


def _write_mape_by_sku(ws: Worksheet, row: int, f: Filters) -> int:
    row = _section_header(ws, row, "Forecast Accuracy by SKU")
    df = query.demand_mape_by_sku(f)
    headers = ["SKU", "Name", "Brand", "MAPE %", "FA %", "Bias %",
               "Volume (cases)", "RAG"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=row, column=i, value=h)
    _apply_header_row(ws, row, len(headers))
    row += 1
    if df.empty:
        ws.cell(row=row, column=1, value="(no data)")
        return row
    for _, r in df.iterrows():
        ws.cell(row=row, column=1, value=r["sku_code"])
        ws.cell(row=row, column=2, value=r["sku_name"])
        brand = r.get("brand")
        ws.cell(row=row, column=3, value=None if pd.isna(brand) else brand)
        for i, (col, fmt) in enumerate(
            (("mape", _FMT_PCT_1DP),
             ("fa_pct", _FMT_PCT_1DP),
             ("bias", _FMT_PCT_SIGNED),
             ("volume", _FMT_INT)),
            start=4,
        ):
            v = r[col]
            cell = ws.cell(
                row=row, column=i,
                value=float(v) if pd.notna(v) else None,
            )
            cell.number_format = fmt
        rag = r["rag"] if pd.notna(r["rag"]) else "red"
        rag_cell = ws.cell(row=row, column=8, value=rag)
        rag_cell.fill = _RAG_FILLS.get(rag, _RAG_FILLS["red"])
        rag_cell.font = Font(
            bold=True, color=config.TKO["colors"]["void"].lstrip("#")
        )
        rag_cell.alignment = Alignment(horizontal="center")
        row += 1
    return row - 1


# ---------------------------------------------------------------------------
# Supply sheet
# ---------------------------------------------------------------------------


def _apply_rag_cell(ws: Worksheet, row: int, col: int, rag: str) -> None:
    cell = ws.cell(row=row, column=col, value=rag)
    cell.fill = _RAG_FILLS.get(rag, _RAG_FILLS["red"])
    cell.font = Font(
        bold=True, color=config.TKO["colors"]["void"].lstrip("#")
    )
    cell.alignment = Alignment(horizontal="center")


def _write_supply_sheet(ws: Worksheet, f: Filters) -> None:
    ws["A1"] = "Supply Review"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A1:F1")
    ws.cell(row=2, column=1,
            value=f"Window {f.period_from:%Y-%m} → {f.period_to:%Y-%m}")

    row = 4
    row = _write_supply_kpi_block(ws, row, f)
    row = _write_supply_dos_by_sku(ws, row + 2, f)
    row = _write_supply_prod_adherence(ws, row + 2, f)
    row = _write_supply_capacity(ws, row + 2, f)
    row = _write_supply_fill_rate_trend(ws, row + 2, f)
    row = _write_supply_gaps(ws, row + 2, f)

    for col, width in zip("ABCDEFGH", (20, 24, 18, 18, 18, 18, 18, 18)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "B2"


def _write_supply_kpi_block(ws: Worksheet, row: int, f: Filters) -> int:
    row = _section_header(ws, row, "Headline KPIs")
    headers = ["Days of Supply", "Fill Rate", "Production Adherence"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=row, column=i, value=h)
    _apply_header_row(ws, row, len(headers))
    row += 1
    k = query.supply_kpis(f)
    values = [k.dos_avg, k.fill_rate, k.production_adherence]
    formats = ['0.0" days"', _FMT_PCT_1DP, _FMT_PCT_1DP]
    for i, (v, fmt) in enumerate(zip(values, formats), start=1):
        if v is None or pd.isna(v):
            ws.cell(row=row, column=i, value=None)
        else:
            cell = ws.cell(row=row, column=i, value=float(v))
            cell.number_format = fmt
    return row


def _write_supply_dos_by_sku(ws: Worksheet, row: int, f: Filters) -> int:
    row = _section_header(ws, row, "DOS by SKU (latest period)")
    df = query.supply_dos_by_sku(f)
    headers = ["SKU", "Name", "DOS (days)", "RAG"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=row, column=i, value=h)
    _apply_header_row(ws, row, len(headers))
    row += 1
    if df.empty:
        ws.cell(row=row, column=1, value="(no data)")
        return row
    for _, r in df.iterrows():
        ws.cell(row=row, column=1, value=r["sku_code"])
        ws.cell(row=row, column=2, value=r["sku_name"])
        dos = r["dos"]
        cell = ws.cell(
            row=row, column=3,
            value=float(dos) if pd.notna(dos) else None,
        )
        cell.number_format = "0.0"
        rag = r["rag"] if pd.notna(r["rag"]) else "red"
        _apply_rag_cell(ws, row, 4, rag)
        row += 1
    return row - 1


def _write_supply_prod_adherence(ws: Worksheet, row: int, f: Filters) -> int:
    row = _section_header(ws, row, "Production Adherence by Plant")
    df = query.supply_production_adherence(f)
    headers = ["Plant", "Name", "Plan (cases)", "Actual (cases)", "Adherence %"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=row, column=i, value=h)
    _apply_header_row(ws, row, len(headers))
    row += 1
    if df.empty:
        ws.cell(row=row, column=1, value="(no data)")
        return row
    for _, r in df.iterrows():
        ws.cell(row=row, column=1, value=r["plant_code"])
        ws.cell(row=row, column=2, value=r["plant_name"])
        for i, (col, fmt) in enumerate(
            (("plan", _FMT_INT),
             ("actual", _FMT_INT),
             ("adherence_pct", _FMT_PCT_1DP)),
            start=3,
        ):
            v = r[col]
            cell = ws.cell(
                row=row, column=i,
                value=float(v) if pd.notna(v) else None,
            )
            cell.number_format = fmt
        row += 1
    return row - 1


def _write_supply_capacity(ws: Worksheet, row: int, f: Filters) -> int:
    row = _section_header(ws, row, "Capacity Utilization by Plant")
    df = query.supply_capacity_utilization(f)
    headers = ["Plant", "Name", "Actual (cases)", "Capacity (cases)", "Utilization %"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=row, column=i, value=h)
    _apply_header_row(ws, row, len(headers))
    row += 1
    if df.empty:
        ws.cell(row=row, column=1, value="(no data)")
        return row
    for _, r in df.iterrows():
        ws.cell(row=row, column=1, value=r["plant_code"])
        ws.cell(row=row, column=2, value=r["plant_name"])
        for i, (col, fmt) in enumerate(
            (("actual", _FMT_INT),
             ("capacity", _FMT_INT),
             ("utilization_pct", _FMT_PCT_1DP)),
            start=3,
        ):
            v = r[col]
            cell = ws.cell(
                row=row, column=i,
                value=float(v) if pd.notna(v) else None,
            )
            cell.number_format = fmt
        row += 1
    return row - 1


def _write_supply_fill_rate_trend(ws: Worksheet, row: int, f: Filters) -> int:
    row = _section_header(ws, row, "Fill Rate Trend (monthly)")
    df = query.supply_fill_rate_trend(f)
    headers = ["Period", "Fill Rate %"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=row, column=i, value=h)
    _apply_header_row(ws, row, len(headers))
    row += 1
    if df.empty:
        ws.cell(row=row, column=1, value="(no data)")
        return row
    for _, r in df.iterrows():
        period = pd.Timestamp(r["period_date"]).date()
        ws.cell(row=row, column=1, value=period).number_format = _FMT_DATE
        v = r["fill_rate"]
        cell = ws.cell(
            row=row, column=2,
            value=float(v) if pd.notna(v) else None,
        )
        cell.number_format = _FMT_PCT_1DP
        row += 1
    return row - 1


def _write_supply_gaps(ws: Worksheet, row: int, f: Filters) -> int:
    row = _section_header(ws, row, "Supply Gaps — projected stockouts (next 3 months)")
    df = query.supply_gaps(f)
    headers = ["SKU", "Name", "Period", "Shortfall (native UOM)"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=row, column=i, value=h)
    _apply_header_row(ws, row, len(headers))
    row += 1
    if df.empty:
        ws.cell(row=row, column=1, value="(no projected shortfalls)")
        return row
    for _, r in df.iterrows():
        ws.cell(row=row, column=1, value=r["sku_code"])
        ws.cell(row=row, column=2, value=r["sku_name"])
        period = pd.Timestamp(r["period_date"]).date()
        ws.cell(row=row, column=3, value=period).number_format = _FMT_DATE
        v = r["shortfall"]
        cell = ws.cell(
            row=row, column=4,
            value=float(v) if pd.notna(v) else None,
        )
        cell.number_format = _FMT_INT
        row += 1
    return row - 1


# ---------------------------------------------------------------------------
# Financial sheet
# ---------------------------------------------------------------------------


def _write_financial_sheet(ws: Worksheet, f: Filters) -> None:
    ws["A1"] = "Financial Review"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A1:F1")
    ws.cell(
        row=2, column=1,
        value=f"Window {f.period_from:%Y-%m} → {f.period_to:%Y-%m} "
              f"· Currency: {config.CURRENCY}",
    )

    row = 4
    row = _write_financial_kpi_block(ws, row, f)
    row = _write_financial_waterfall(ws, row + 2, f)
    row = _write_financial_by_channel(ws, row + 2, f)
    row = _write_financial_pnl_summary(ws, row + 2, f)
    row = _write_financial_ytd_progress(ws, row + 2, f)
    row = _write_financial_gm_trend(ws, row + 2, f)

    for col, width in zip("ABCDEFGH", (20, 24, 22, 22, 22, 18, 18, 18)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "B2"


def _write_financial_kpi_block(ws: Worksheet, row: int, f: Filters) -> int:
    row = _section_header(ws, row, "Headline KPIs")
    headers = [
        f"Revenue Actual ({config.CURRENCY})",
        "Revenue vs Budget",
        "Revenue vs LE",
        "GM %",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=row, column=i, value=h)
    _apply_header_row(ws, row, len(headers))
    row += 1
    k = query.financial_kpis(f)
    values  = [k.revenue_actual, k.revenue_vs_budget_pct, k.revenue_vs_le_pct, k.gm_pct]
    formats = [_FMT_INT, _FMT_PCT_SIGNED, _FMT_PCT_SIGNED, _FMT_PCT_1DP]
    for i, (v, fmt) in enumerate(zip(values, formats), start=1):
        if v is None or pd.isna(v):
            ws.cell(row=row, column=i, value=None)
        else:
            cell = ws.cell(row=row, column=i, value=float(v))
            cell.number_format = fmt
    return row


def _write_financial_waterfall(ws: Worksheet, row: int, f: Filters) -> int:
    row = _section_header(ws, row, "Revenue Waterfall (Budget → LE → Actuals)")
    df = query.financial_revenue_waterfall(f)
    headers = ["Stage", f"Revenue ({config.CURRENCY})"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=row, column=i, value=h)
    _apply_header_row(ws, row, len(headers))
    row += 1
    if df.empty:
        ws.cell(row=row, column=1, value="(no data)")
        return row
    for _, r in df.iterrows():
        ws.cell(row=row, column=1, value=r["stage"])
        cell = ws.cell(row=row, column=2, value=float(r["value"]))
        cell.number_format = _FMT_INT
        row += 1
    return row - 1


def _write_financial_by_channel(ws: Worksheet, row: int, f: Filters) -> int:
    row = _section_header(ws, row, "Revenue by Channel")
    df = query.financial_by_channel(f)
    headers = ["Channel Code", "Channel", f"Revenue ({config.CURRENCY})"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=row, column=i, value=h)
    _apply_header_row(ws, row, len(headers))
    row += 1
    if df.empty:
        ws.cell(row=row, column=1, value="(no data)")
        return row
    for _, r in df.iterrows():
        ws.cell(row=row, column=1, value=r["channel_code"])
        ws.cell(row=row, column=2, value=r["channel_name"])
        v = r["revenue_actual"]
        cell = ws.cell(
            row=row, column=3,
            value=float(v) if pd.notna(v) else None,
        )
        cell.number_format = _FMT_INT
        row += 1
    return row - 1


def _write_financial_pnl_summary(ws: Worksheet, row: int, f: Filters) -> int:
    row = _section_header(ws, row, "P&L Summary by Brand")
    df = query.financial_pnl_summary(f)
    headers = [
        "Brand",
        f"Revenue ({config.CURRENCY})",
        f"GM ({config.CURRENCY})",
        f"Promo Spend ({config.CURRENCY})",
        f"Net Revenue ({config.CURRENCY})",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=row, column=i, value=h)
    _apply_header_row(ws, row, len(headers))
    row += 1
    if df.empty:
        ws.cell(row=row, column=1, value="(no data)")
        return row
    for _, r in df.iterrows():
        brand = r["brand"]
        ws.cell(row=row, column=1, value=None if pd.isna(brand) else brand)
        for i, col in enumerate(
            ("revenue", "gm", "promo_spend", "net_revenue"), start=2
        ):
            v = r[col]
            cell = ws.cell(
                row=row, column=i,
                value=float(v) if pd.notna(v) else None,
            )
            cell.number_format = _FMT_INT
        row += 1
    return row - 1


def _write_financial_ytd_progress(ws: Worksheet, row: int, f: Filters) -> int:
    row = _section_header(
        ws, row, f"YTD vs Full-Year Budget ({f.period_to.year})"
    )
    df = query.financial_ytd_progress(f)
    headers = [
        "Brand",
        f"YTD Actual ({config.CURRENCY})",
        f"Full-Year Budget ({config.CURRENCY})",
        "YTD %",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=row, column=i, value=h)
    _apply_header_row(ws, row, len(headers))
    row += 1
    if df.empty:
        ws.cell(row=row, column=1, value="(no data)")
        return row
    for _, r in df.iterrows():
        brand = r["brand"]
        ws.cell(row=row, column=1, value=None if pd.isna(brand) else brand)
        for i, (col, fmt) in enumerate(
            (("ytd_actual", _FMT_INT),
             ("full_year_budget", _FMT_INT),
             ("ytd_pct", _FMT_PCT_1DP)),
            start=2,
        ):
            v = r[col]
            cell = ws.cell(
                row=row, column=i,
                value=float(v) if pd.notna(v) else None,
            )
            cell.number_format = fmt
        row += 1
    return row - 1


def _write_financial_gm_trend(ws: Worksheet, row: int, f: Filters) -> int:
    row = _section_header(ws, row, "GM% Trend (monthly)")
    df = query.financial_gm_trend(f)
    headers = ["Period", "GM %"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=row, column=i, value=h)
    _apply_header_row(ws, row, len(headers))
    row += 1
    if df.empty:
        ws.cell(row=row, column=1, value="(no data)")
        return row
    for _, r in df.iterrows():
        period = pd.Timestamp(r["period_date"]).date()
        ws.cell(row=row, column=1, value=period).number_format = _FMT_DATE
        v = r["gm_pct"]
        cell = ws.cell(
            row=row, column=2,
            value=float(v) if pd.notna(v) else None,
        )
        cell.number_format = _FMT_PCT_1DP
        row += 1
    return row - 1


# ---------------------------------------------------------------------------
# Scenario sheet (Phase 7)
# ---------------------------------------------------------------------------
# Volume × current avg-price proxy per (sku, channel) per scenario. SPEC §5.5:
# user enters one % adjustment per scenario; that % is applied uniformly to the
# forward consensus volume to get scenario volume; revenue = volume × price.


def _write_scenario_sheet(
    ws: Worksheet, f: Filters, adjustments: dict
) -> None:
    ws["A1"] = "Scenario Comparison"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A1:F1")
    ws.cell(
        row=2, column=1,
        value=f"Window {f.period_from:%Y-%m} → {f.period_to:%Y-%m} "
              f"· Currency: {config.CURRENCY}",
    )

    summary = _scenario_summary(f, adjustments)

    row = 4
    row = _section_header(ws, row, "Scenario Headline")
    headers = [
        "Scenario", "% vs Consensus", "Volume (cases)",
        f"Revenue ({config.CURRENCY})", "Δ Revenue vs Base",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=row, column=i, value=h)
    _apply_header_row(ws, row, len(headers))
    row += 1

    if summary.empty:
        ws.cell(
            row=row, column=1,
            value="(no forward planning horizon — actuals extend to "
                  "the end of the window; scenarios have no months to project)",
        )
    else:
        base_revenue = float(
            summary.loc[summary["scenario"] == "Base", "revenue"].iloc[0]
        ) if (summary["scenario"] == "Base").any() else None
        for _, r in summary.iterrows():
            ws.cell(row=row, column=1, value=r["scenario"])
            adj_cell = ws.cell(row=row, column=2, value=float(r["adjustment_pct"]))
            adj_cell.number_format = _FMT_PCT_SIGNED
            vol_cell = ws.cell(row=row, column=3, value=float(r["volume_cases"]))
            vol_cell.number_format = _FMT_INT
            rev_cell = ws.cell(row=row, column=4, value=float(r["revenue"]))
            rev_cell.number_format = _FMT_INT
            if base_revenue is None or r["scenario"] == "Base":
                ws.cell(row=row, column=5, value=None)
            else:
                delta_cell = ws.cell(
                    row=row, column=5,
                    value=float(r["revenue"] - base_revenue),
                )
                delta_cell.number_format = '+#,##0;-#,##0;0'
            row += 1

    # Inputs block
    row += 2
    row = _section_header(ws, row, "Scenario Inputs")
    for label_key in ("base", "upside", "downside"):
        ws.cell(row=row, column=1, value=label_key.capitalize()).font = _LABEL_FONT
        cell = ws.cell(row=row, column=2, value=float(adjustments.get(label_key, 0.0)))
        cell.number_format = _FMT_PCT_SIGNED
        row += 1

    # Methodology footer — keeps the sheet self-explanatory for a recipient
    # who never saw the dashboard.
    row += 1
    ws.cell(
        row=row, column=1,
        value=(
            "Method: volume = consensus_fcst × (1 + %adj) over the FORWARD "
            "portion of the window (after the latest actuals period), "
            "UOM-normalized to cases. Revenue = volume × trailing-3M avg price "
            "per (SKU, channel)."
        ),
    )

    for col, width in zip("ABCDE", (16, 18, 22, 22, 22)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "B2"


def _scenario_summary(f: Filters, adjustments: dict) -> pd.DataFrame:
    """Return per-scenario {volume_cases, revenue} aggregated across (sku, channel).

    Pairs missing a price proxy contribute volume but not revenue — matches
    the page banner's "X pairs excluded" disclosure.
    """
    base = query.scenario_baseline(f)
    price = query.scenario_price_proxy(f)
    if base.empty:
        return pd.DataFrame(
            columns=["scenario", "adjustment_pct", "volume_cases", "revenue"]
        )

    merged = base.merge(price, on=["sku_code", "channel_code"], how="left")
    rows = []
    for name in ("Base", "Upside", "Downside"):
        adj = float(adjustments.get(name.lower(), 0.0))
        vol = merged["consensus_volume_cases"] * (1.0 + adj / 100.0)
        rev = (vol * merged["avg_price"]).fillna(0.0)
        rows.append({
            "scenario": name,
            "adjustment_pct": adj,
            "volume_cases": float(vol.sum()),
            "revenue": float(rev.sum()),
        })
    return pd.DataFrame(rows)
